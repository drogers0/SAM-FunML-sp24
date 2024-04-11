import sys, os
import cv2
import statistics
import time
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backend_bases import MouseButton
from openpyxl import Workbook, load_workbook
from segment_anything import sam_model_registry, SamPredictor

from config import *
from helpers import *

sys.path.append("..")

try:
    matplotlib.use('Qt5Agg')
except:
    matplotlib.use('TkAgg')

plt.rcParams['keymap.grid'].remove('g')
plt.rcParams['keymap.home'].remove('r')

sam = sam_model_registry[MODEL_TYPE](checkpoint=SAM_CHECKPOINT)
sam.to(device=DEVICE)

predictor = SamPredictor(sam)

names = np.load("samples.npy", allow_pickle=True)
labels = np.load("labels.npy", allow_pickle=True)

# %%

first = input("Do you want to load previous work? -y -n\n")
while first != 'n' and first != 'y':
    first = input("Chose y or n, Do you want to load previous work? -y -n\n")

name = input("what is your name?\n")
if first == 'n':
    wb, ws = create_workbook(name)
    c      = 0
    tim    = 0
    serv=np.array([])
else:
    wb = load_workbook(os.path.join(name, name + ".xlsx"))
    ws = wb.active
    c = len(os.listdir(os.path.join(name, "masks")))
    f = open(os.path.join(name, "time.txt"), 'r')
    serv=np.load(os.path.join(name,"servey.npy")) if os.path.exists(os.path.join(name,"servey.npy")) else np.array([])
    tim = f.readline()
    f.close()

t = time.time()



f = False
## start looping through samples: 
while c < 400 and not f:
    msk = []  # masks for each samples

    gp = []  # green points
    rp = []  # red points
    image = names[c]  # samples c
    ws['A' + str(c + 2)] = str(c)  # samples name on excel
    if len(image.shape) == 2:
        image = cv2.cvtColor((np.array(((image + 1) / 2) * 255, dtype='uint8')), cv2.COLOR_GRAY2RGB)
    label = labels[c]  # GT for sample c
    rmv = False
    mask = 0

    predictor.set_image(image)
    inc = ""
    co = 0
    bs = 0
    score = []
    round=[0,0]
    stdx = []
    stdy = []
    ng = []
    nr = []
    green = []
    red = []
    greenx = []
    redx = []
    greeny = []
    redy = []
    label = label == 1



    while inc != "y":
        s = 0  # this is for the score
        count = 1  # to count the score max
        lessfive = 0
        current_color = 'green'
        dot_size_toggle = SMALL_DOT_SIZE_MODE # default will be small dot, not medium
        current_star_size = SMALL_STAR_SIZE
        current_green_red_dot_size = SMALL_GREEN_RED_DOT_SIZE
        # get_ipython().run_line_magic('matplotlib', 'qt')
        fig, ax = plt.subplots(1, 3, figsize=(15, 7))
        if green and red:
            ax[0].plot(greenx, greeny, 'go', markersize=5)
            ax[1].plot(greenx, greeny, 'go', markersize=5)
            ax[0].plot(redx, redy, 'ro', markersize=5)
            ax[1].plot(redx, redy, 'ro', markersize=5)
            plt.draw()


        def onclose(event):
            fig.canvas.stop_event_loop()
            fig.canvas.mpl_disconnect(cid)


        def onclick(event):
            global count
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global label
            global mask
            global lessfive
            if event.xdata is not None and event.ydata is not None:

                x, y = int(event.xdata), int(event.ydata)
                print(not x)
                print(not y)

                if event.button is MouseButton.LEFT:
                    if current_color == 'green':

                        green.append((x, y))
                        greenx.append(x)

                        greeny.append(y)
                        ax[0].plot(x, y, 'go', markersize=current_green_red_dot_size, color = GREEN_COLOR)
                        ax[1].plot(x, y, 'go', markersize=current_green_red_dot_size, color = GREEN_COLOR)
                        plt.draw()

                    else:
                        red.append((x, y))
                        redx.append(x)

                        redy.append(y)
                        ax[0].plot(x, y, 'ro', markersize=current_green_red_dot_size, color = RED_COLOR)
                        ax[1].plot(x, y, 'ro', markersize=current_green_red_dot_size, color = RED_COLOR)
                        plt.draw()

                elif event.button is MouseButton.RIGHT:

                    if not green and not red:
                        print("no points to delete")
                    elif green:
                        print(current_color)
                        if current_color == 'green':
                            indx = closetn((x, y), green)
                            print(indx)
                            for line in ax[0].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == green[indx][0] and line.get_ydata()[0] == green[indx][1]:

                                        line.set_data([], [])
                                        break
                            for line in ax[1].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == green[indx][0] and line.get_ydata()[0] == green[indx][1]:
                                        line.set_data([], [])
                                        break
                            del green[indx]
                            del greenx[indx]

                            del greeny[indx]

                            plt.draw()
                        elif red:
                            print("delete red")
                            print(current_color)
                            indx = closetn((x, y), red)
                            print(indx)

                            for line in ax[0].lines:
                                if len(line.get_xdata()) > 0:
                                    print()
                                    if line.get_xdata()[0] == red[indx][0] and line.get_ydata()[0] == red[indx][1]:
                                        line.set_data([], [])
                                        break
                            for line in ax[1].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == red[indx][0] and line.get_ydata()[0] == red[indx][1]:
                                        line.set_data([], [])
                                        break

                            del red[indx]
                            del redx[indx]

                            del redy[indx]
                            plt.draw()

                if green and red:
                    global s
                    print("green:", green)
                    print("red:", red)

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))

                    masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                    )

                    mask = masks[0]

                    ax[2].clear()
                    ax[2].imshow(image)
                    show_mask(mask, ax[2])
                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union

                    show_points(input_point, input_label, ax[2], marker_size = current_star_size)
                    msg = ""

                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx",maxx)
                    score.append(s)
                    gp.append(np.multiply(green, 1))

                    rp.append(np.multiply(red, 1))
                    ng.append(len(greenx))
                    nr.append(len(redx))
                    grx = np.concatenate([greenx, redx])
                    gry = np.concatenate([greeny, redy])

                    stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                    stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                    print("up count", count)
                    if maxx >= s:
                        print("inside",count)
                        if count >= 10:
                            lessfive += 1
                        else:
                            count += 1
                    elif maxx < s:

                        count = 1
                    if lessfive == 1:
                        maxx = 0
                        count=1
                        round[0] = len(np.array(score))
                        msg = " (round 2) "
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))
                    print("less than best score", lessfive)
                    print("scores:", score)
                    if lessfive == 1:
                        lessfive += 1
                        for line in ax[0].lines:
                            line.set_data([], [])
                        for line in ax[1].lines:
                            line.set_data([], [])
                        green = []
                        red = []
                        greenx = []
                        redx = []
                        greeny = []
                        redy = []
                        plt.draw()
                        ax[2].clear()
                        ax[2].imshow(image)
                        show_mask(mask, ax[2])
                        count = 1
                        print("below count", count)
                        plt.title("No better score is achieved in the last 5 attempts. Start round 2 from scratch")
                    elif lessfive == 3:
                        round[1]=len(score)-round[0]
                        print("The window closed because you did not achieve a better score after 5 consecutive clicks in the 2nd round")
                        plt.close()


        # Create a function to toggle between green and red dots
        def toggle_color(event):
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global current_color
            global count
            global current_star_size
            global current_green_red_dot_size
            global dot_size_toggle
            
            if event.key == 'g':
                current_color = 'green'
                print("Switched to GREEN dot mode.")

            elif event.key == 'r':
                current_color = 'red'
                print("Switched to RED dot mode.")
            elif event.key == ' ':
                for line in ax[0].lines:
                    line.set_data([], [])
                for line in ax[1].lines:
                    line.set_data([], [])
                green = []
                red = []
                greenx = []
                redx = []
                greeny = []
                redy = []
                plt.draw()
                ax[2].clear()
                ax[2].imshow(image)
                show_mask(mask, ax[2])
                count = 1
                print("below count", count)
            elif event.key == 'z':
                dot_size_toggle = not dot_size_toggle
                
                if dot_size_toggle == SMALL_DOT_SIZE_MODE:
                    # true => smaller dot size
                    current_star_size = SMALL_STAR_SIZE
                    current_green_red_dot_size = SMALL_GREEN_RED_DOT_SIZE
                    print("Switched to SMALL DOT SIZE mode.")
                else:
                    # false => default dot size
                    current_star_size = MEDIUM_STAR_SIZE
                    current_green_red_dot_size = MEDIUM_GREEN_RED_DOT_SIZE
                    print("Switched to MEDIUM DOT SIZE mode.")
                
                

        # Create a figure and display the image

        a = ax[0].plot()
        b = ax[1].plot()
        ax[0].imshow(image)
        ax[1].imshow(label)
        # Connect mouse click and keyboard key events
        fig.canvas.mpl_connect('button_press_event', onclick)
        # fig.canvas.start_event_loop(timeout=-5)
        fig.canvas.mpl_connect('key_press_event', toggle_color)
        fig.canvas.mpl_connect('key_press_event', toggle_color)
        # fig.canvas.start_event_loop(timeout=-5)
        # Display the plot

        cid = fig.canvas.mpl_connect('close_event', onclose)
        fig.show()  # this call does not block on my system
        fig.canvas.start_event_loop()  # block here until window closed

        inc = "y"
        print(inc)

    indx = np.argsort(-np.array(score))
    sscore = np.array(score)[indx]
    sng = np.array(ng)[indx]
    snr = np.array(nr)[indx]
    sstdx = np.array(stdx)[indx]
    sstdy = np.array(stdy)[indx]
    for i in range(len(score)):
        coun = 1
        for col in ws.iter_cols(min_row=c + 2, max_row=c + 2, max_col=6 + i * 5, min_col=2 + i * 5):
            if coun == 1:
                ws[col[0].coordinate] = sng[i]
            elif coun == 2:
                ws[col[0].coordinate] = snr[i]
            elif coun == 3:
                ws[col[0].coordinate] = sstdx[i]
            elif coun == 4:
                ws[col[0].coordinate] = sstdy[i]
            elif coun == 5:
                ws[col[0].coordinate] = sscore[i]
            coun += 1
    np.save(os.path.join(name, "points", str(c) + "_green"), np.array(gp, dtype=object))
    np.save(os.path.join(name, "points", str(c) + "_red"), np.array(rp, dtype=object))
    np.save(os.path.join(name, "masks", str(c) + "_mask"), np.array(msk))
    np.save(os.path.join(name, "sorts", str(c) + "_sort"), indx)
    np.save(os.path.join(name, "scores", str(c) + "score"), score)
    np.save(os.path.join(name,"eachround",str(c)+"_"),round)

    c += 1
    ans=input("Do you think the ground truth mask was suboptimal? (i.e. are SAM's results qualitatively better) y or n\n") 
    while ans!="y" and ans!="n":
        ans=input("Do you think the ground truth mask was suboptimal? (i.e. are SAM's results qualitatively better) y or n\n") 
    ans = 1 if ans=="y" else 0 
    
    serv=np.append(serv,ans)
    contin = input("do u want to continue? press y if you want to continue or anyting otherwise ")
    if not contin == 'y':
        wb.save(os.path.join(name, name + '.xlsx'))
        f = True


    print("Sample:", c)
wb.save(os.path.join(name, name + '.xlsx'))
file = open(os.path.join(name, "time.txt"), 'w')
file.write(str(float(tim) + (time.time() - t)))
np.save(os.path.join(name,"servey.npy"),serv)
file.close()
